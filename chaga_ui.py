import tkinter as tk
from tkinter import messagebox
import psycopg2
import openpyxl
from openpyxl.styles import Font

def salvar_item():
    nome = entry_nome.get()
    categoria = entry_categoria.get()
    preco = entry_preco.get()
    quantidade = entry_quantidade.get()

    if not nome or not preco or not quantidade:
        messagebox.showwarning("Erro", "Preencha todos os campos obrigatórios")
        return
    
    try:
        preco_val = float(preco)
        quantidade_val = int(quantidade)

    except ValueError:
        messagebox.showerror("Erro", "Preço deve ser número e Quantidade deve ser inteiro")
        return
    
    try:
        conn = psycopg2.connect(
          dbname = "loja",
          user= "postgres",
          password= "pg@10",
          host="localhost",
         port= "5432"
        )

        cur = conn.cursor()
        cur.execute(""" 
          INSERT INTO itens (nome, categoria, preco, quantidade)
          VALUES (%s, %s, %s, %s)
         """, (nome, categoria, preco_val, quantidade_val))
        conn.commit()
        cur.close()
        conn.close()

        messagebox.showinfo("Sucesso", "Item cadastrado!")
        entry_nome.delete(0,  tk.END)
        entry_categoria.delete(0, tk.END)
        entry_preco.delete(0, tk.END)
        entry_quantidade.delete(0, tk.END)

        gerar_planilha()

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar no banco:\n{e}")



def deletar_item():
    nome = entry_nome.get()

    if not nome:
        messagebox.showwarning("Erro", "Digite o nome do item para deletar")
        return
    try:
        conn = psycopg2.connect(
        dbname = "loja",
        user= "postgres",
        password= "pg@10",
        host="localhost",
        port= "5432"
     )
    
        cur = conn.cursor()
        cur.execute("DELETE FROM itens WHERE nome= %s", (nome,))
        conn.commit()
        cur.close()
        conn.close()

        messagebox.showinfo("Sucesso", f"Item '{nome}' deletado.")
        entry_nome.delete(0, tk.END)
        entry_preco.delete(0, tk.END)
        entry_quantidade.delete(0, tk.END)

        gerar_planilha()

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao deletar:\n{e}")


def dar_baixa():
    nome = entry_nome.get()
    qtd_baixa = entry_baixa.get()

    if not nome:
        messagebox.showwarning("Erro", "Digite o nome e a quantidade para dar baixa")
        return
    
    try:
        qtd_baixa_val = int(qtd_baixa)
        if qtd_baixa_val <= 0:
            raise ValueError("Deve inserir ao menos um valor positivo")
        
    except ValueError:
        messagebox.showerror("Erro", "quantidade para dar baixa deve ser um número inteiro positivo")
        return

    try:
        conn = psycopg2.connect(
        dbname = "loja",
        user= "postgres",
        password= "pg@10",
        host="localhost",
        port= "5432"
     )
        
        cur = conn.cursor()
        cur.execute("SELECT quantidade FROM itens WHERE nome = %s", (nome,))
        resultado = cur.fetchone()


        if not resultado:
            messagebox.showerror("Erro", "Item não encontrado")
        else:
            quantidade_atual = resultado[0]
            if quantidade_atual < qtd_baixa_val:
                messagebox.showwarning("Aviso", "Estoque insuficiente. Apenas {quantidade_atual} disponíveis")
            else:
                cur.execute(
                    "UPDATE itens SET quantidade = quantidade - %s WHERE nome = %s",(qtd_baixa_val,nome,))
                conn.commit()

                messagebox.showinfo("Sucesso", f"Baixa realizada. Novo estoque: {quantidade_atual - qtd_baixa_val}")
        cur.close()
        conn.close()
       

        entry_nome.delete(0, tk.END)
        entry_categoria.delete(0, tk.END)
        entry_preco.delete(0, tk.END)
        entry_quantidade.delete(0, tk.END)

        gerar_planilha()

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao dar baixa:\n{e}")    

def gerar_planilha():
    try:
        conn = psycopg2.connect(
        dbname = "loja",
        user= "postgres",
        password= "pg@10",
        host="localhost",
        port= "5432"
     )
        cur = conn.cursor()
        cur.execute("SELECT nome, categoria, preco, quantidade FROM itens ORDER BY nome")
        dados = cur.fetchall()
     
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estoque"

 
        cabecalhos = ["nome", "Categoria", "Preço", "Quantidade"]
        for col, cab in enumerate(cabecalhos, start=1):
          cell = ws.cell(row=1, column=col, value=cab)
          cell.font = Font(bold= True)


        for i, linha in enumerate(dados, start = 2):
           for j, valor in enumerate(linha, start=1):
             ws.cell(row=i, column=j, value=valor)

        wb.save("..:\\...\\..\\..\\...\\...\\estoque.xlsx")
        cur.close()
        conn.close()

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao gerar a planilha:\n{e}")
 




janela = tk.Tk()
janela.title("Itens do Chaga")
janela.geometry("500x400")

frame = tk.Frame(janela)
frame.pack(expand=True)

tk.Label(frame, text="Nome:").grid(row=0, column= 0, padx=5, pady=5, sticky="e")
entry_nome = tk.Entry(frame, width=30)
entry_nome.grid(row=0, column=1, pady=5)

tk.Label(frame, text="Categoria").grid(row=1, column= 0, padx=5, pady=5, sticky="e")
entry_categoria = tk.Entry(frame, width=30)
entry_categoria.grid(row=1, column=1, pady=5)


tk.Label(frame, text="Preço").grid(row=2, column= 0, padx=5, pady=5, sticky="e")
entry_preco = tk.Entry(frame, width=30)
entry_preco.grid(row=2, column=1, pady=5)


tk.Label(frame, text="Quantidade").grid(row=3, column= 0, padx=5, pady=5, sticky="e")
entry_quantidade = tk.Entry(frame, width=30)
entry_quantidade.grid(row=3, column=1, pady=5)

tk.Label(frame, text="Dar baixa (Qtd)").grid(row=4, column= 0, padx=5, pady=5, sticky="e")
entry_baixa = tk.Entry(frame, width=30)
entry_baixa.grid(row=4, column=1, pady=5)


btn_salvar = tk.Button(frame, text="Salvar", command=salvar_item)
btn_salvar.grid(row=5, column=0, columnspan=2, pady=30)

btn_deletar= tk.Button(frame, text="Deletar", command=deletar_item)
btn_deletar.grid(row=5, column=1, pady=5)

btn_baixa = tk.Button(frame, text="Dar baixa (vendido)", command=dar_baixa)
btn_baixa.grid(row=5, column=2, pady=5)

janela.mainloop()





